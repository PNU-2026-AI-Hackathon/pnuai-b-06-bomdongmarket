"""0.3.1 핵심 계산과 Excel 출력 회귀 테스트."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from excel_output import write_profit_output
from main import calculate_all_sites
from profit_calculation import (
    LONG_TERM_RECOMMENDATION,
    SHORT_TERM_RECOMMENDATION,
    calculate_profit,
)


class ProfitCalculationTest(unittest.TestCase):
    def test_all_spaces_are_combined_with_all_crops(self) -> None:
        scenarios = calculate_all_sites()

        self.assertEqual(len(scenarios), 9)
        by_site: dict[str, set[str]] = {}
        for scenario in scenarios:
            meta = scenario["space_row"]
            self.assertIsInstance(meta, dict)
            by_site.setdefault(str(meta["site_id"]), set()).add(
                str(meta["crop_name"])
            )

        self.assertEqual(set(by_site), {"S001", "S002", "S003"})
        for crops in by_site.values():
            self.assertEqual(crops, {"상추", "딸기", "바질"})

    def test_equal_expected_income_recommends_long_term(self) -> None:
        result = calculate_profit(
            monthly_revenue_krw=1_000,
            monthly_electricity_cost_krw=0,
            monthly_water_cost_krw=0,
            monthly_material_cost_krw=0,
            monthly_labor_cost_krw=0,
            desired_monthly_rent_krw=800,
            standard={"depreciation_and_other_cost_krw_month": 0},
            contract={"landlord_share_ratio": 0.8},
        )

        self.assertEqual(result["landlord_expected_income_krw"], 800)
        self.assertEqual(result["recommendation"], LONG_TERM_RECOMMENDATION)
        self.assertEqual(result["contract_type"], "장기계약형")

    def test_operating_loss_is_preserved_and_recommends_short_term(self) -> None:
        result = calculate_profit(
            monthly_revenue_krw=100,
            monthly_electricity_cost_krw=200,
            monthly_water_cost_krw=0,
            monthly_material_cost_krw=0,
            monthly_labor_cost_krw=0,
            desired_monthly_rent_krw=0,
            standard={"depreciation_and_other_cost_krw_month": 0},
            contract={"landlord_share_ratio": 0.8},
        )

        self.assertEqual(result["monthly_operating_profit_krw"], -100)
        self.assertEqual(result["landlord_expected_income_krw"], -80)
        self.assertEqual(result["business_operating_profit_krw"], -20)
        self.assertEqual(result["recommendation"], SHORT_TERM_RECOMMENDATION)
        self.assertEqual(result["contract_type"], "단기계약형")

    def test_excel_contains_nine_scenarios_and_required_sheets(self) -> None:
        scenarios = calculate_all_sites()
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "Profit_Output.xlsx"
            write_profit_output(scenarios, output_path)
            workbook = load_workbook(output_path, data_only=False)

            self.assertEqual(
                workbook.sheetnames,
                ["요약", "계산상세", "월별전력량", "입력기준", "검증"],
            )
            self.assertEqual(workbook["요약"].max_row, 13)
            self.assertEqual(workbook["요약"]["A5"].value, "S001-상추")
            self.assertEqual(workbook["요약"]["A13"].value, "S003-바질")
            self.assertEqual(workbook["월별전력량"].max_row, 111)


if __name__ == "__main__":
    unittest.main()
